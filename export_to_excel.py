from pathlib import Path
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


DATA_DIR = Path("data")
CSV_FILE = DATA_DIR / "pink_sheet.csv"
EXCEL_FILE = DATA_DIR / "pink_sheet.xlsx"


def main():
    df = pd.read_csv(CSV_FILE)

    with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Pink Sheet")

        ws = writer.book["Pink Sheet"]

        light_pink = PatternFill(fill_type="solid", fgColor="FCE4EC")
        white = PatternFill(fill_type="solid", fgColor="FFFFFF")
        thin_border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )

        # Fixed column width + alternate column colours
        for col_idx in range(1, ws.max_column + 1):
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            ws.column_dimensions[col_letter].width = 10

            fill = light_pink if col_idx % 2 == 0 else white

            for row_idx in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = fill
                cell.font = Font(name="Arial Narrow", size=10, bold=False)
                cell.alignment = Alignment(
                    wrap_text=True,
                    horizontal="center",
                    vertical="center"
                )
                cell.border = thin_border

        # Header row only
        ws.row_dimensions[1].height = 60

        for cell in ws[1]:
            cell.font = Font(name="Arial Narrow", size=10, bold=False)
            cell.alignment = Alignment(
                wrap_text=True,
                horizontal="center",
                vertical="center"
            )

        ws.freeze_panes = "A2"

    print("Pink Sheet Excel formatting applied successfully.")


if __name__ == "__main__":
    main()
